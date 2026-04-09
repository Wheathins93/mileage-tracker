"""
Mileage Tracker — Flask Application

Endpoints:
  - /api/auth/*       — PIN-based registration, login, verification
  - /api/branches     — Branch list
  - /api/routes       — Routes between branches
  - /api/entries      — CRUD for mileage entries (scoped to user)
  - /api/export/*     — CSV and Excel export
"""

import copy
import csv
import io
import os
import calendar
from datetime import datetime
from functools import wraps

from flask import Flask, render_template, request, jsonify, Response

from openpyxl import load_workbook

from database import get_db, init_db, create_user, authenticate_user
from routes_data import (
    BRANCHES,
    MILEAGE_TABLE,
    REIMBURSEMENT_RATE,
    get_routes,
    get_route_miles,
    calculate_reimbursement,
)

app = Flask(__name__)

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "Blank Expense Reimbursement Form - 2026.xlsx",
)


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------
with app.app_context():
    init_db()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def get_user_id():
    """Read the user_id from the X-User-Id header set by the frontend."""
    return request.headers.get("X-User-Id", "").strip()


def require_user(f):
    """Decorator that returns 401 if no user_id header is present."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not get_user_id():
            return jsonify({"error": "Not authenticated. Please log in."}), 401
        return f(*args, **kwargs)

    return decorated


def _validate_entry(data):
    """Validate incoming entry data. Returns list of error strings."""
    errors = []
    if not data:
        return ["No data provided"]
    for field in (
        "year", "month", "day",
        "origin_branch", "destination_branch", "route_name",
    ):
        if not data.get(field):
            errors.append(f"'{field}' is required")
    if data.get("origin_branch") == data.get("destination_branch"):
        errors.append("Origin and destination must be different")
    return errors


def _build_entry_values(data, user_id, *, override_origin=None,
                        override_dest=None, override_route=None):
    """Build the tuple of values for an INSERT.

    Optional overrides let us reuse this for the return-trip entry.
    Returns (values_tuple, error_string_or_None).
    """
    origin = override_origin or data["origin_branch"]
    dest = override_dest or data["destination_branch"]
    route_name = override_route or data["route_name"]

    miles = get_route_miles(origin, dest, route_name)
    if miles is None:
        return None, "Invalid route selection"

    reimbursement = calculate_reimbursement(miles)
    year = int(data["year"])
    month = int(data["month"])
    day = int(data["day"])
    date_str = f"{year}-{month:02d}-{day:02d}"

    purpose = data.get("business_purpose", "").strip()
    if not purpose or override_origin:
        purpose = f"Mileage: {origin} to {dest} via {route_name}"

    notes = data.get("notes", "").strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return (
        user_id, year, month, day, date_str,
        origin, dest, route_name,
        miles, reimbursement, purpose, notes, now, now,
    ), None


INSERT_SQL = """
    INSERT INTO entries
        (user_id, year, month, day, date, origin_branch, destination_branch,
         route_name, miles, reimbursement_amount, business_purpose, notes,
         created_at, updated_at)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# API — Authentication
# ---------------------------------------------------------------------------
@app.route("/api/auth/register", methods=["POST"])
def api_register():
    """Register a new user with a PIN and display name."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    pin = data.get("pin", "").strip()
    display_name = data.get("display_name", "").strip()

    if not pin or not display_name:
        return jsonify({"error": "PIN and name are required"}), 400

    if not pin.isdigit() or len(pin) < 4 or len(pin) > 8:
        return jsonify({"error": "PIN must be 4–8 digits"}), 400

    user = create_user(pin, display_name)
    if user is None:
        return jsonify({
            "error": "That PIN is already in use. Please choose a different one."
        }), 409

    return jsonify({
        "user_id": user["id"],
        "display_name": user["display_name"],
    }), 201


@app.route("/api/auth/login", methods=["POST"])
def api_login():
    """Log in with a PIN."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400

    pin = data.get("pin", "").strip()
    if not pin:
        return jsonify({"error": "PIN is required"}), 400

    user = authenticate_user(pin)
    if user is None:
        # Check if ANY users exist to give a more helpful error message
        from database import get_user_count
        count = get_user_count()
        if count == 0:
            return jsonify({
                "error": "No accounts exist yet. Please create an account first."
            }), 401
        return jsonify({
            "error": "No account found for that PIN. Tap 'Create an account' to register."
        }), 401

    return jsonify({
        "user_id": user["id"],
        "display_name": user["display_name"],
    })


@app.route("/api/auth/verify", methods=["POST"])
def api_verify():
    """Verify that a stored user_id is still valid.

    Also returns ``user_count`` so the frontend can decide whether to
    show the login form or the register form.
    """
    data = request.get_json()
    if not data:
        return jsonify({"valid": False, "user_count": 0}), 400

    user_id = data.get("user_id", "").strip()
    if not user_id:
        return jsonify({"valid": False, "user_count": 0}), 400

    conn = get_db()
    try:
        user = conn.execute(
            "SELECT id, display_name FROM users WHERE id = ?", (user_id,)
        ).fetchone()
        count_row = conn.execute("SELECT COUNT(*) as cnt FROM users").fetchone()
        user_count = count_row["cnt"] if count_row else 0
    finally:
        conn.close()

    if user:
        return jsonify({
            "valid": True,
            "user_id": user["id"],
            "display_name": user["display_name"],
            "user_count": user_count,
        })
    return jsonify({"valid": False, "user_count": user_count}), 401


@app.route("/api/auth/status")
def api_auth_status():
    """Check if any registered users exist.

    Used on first load (no stored session) to decide whether to show
    the login form or the registration form.
    """
    from database import get_user_count
    count = get_user_count()
    return jsonify({"user_count": count})


# ---------------------------------------------------------------------------
# API — Reference Data
# ---------------------------------------------------------------------------
@app.route("/api/branches")
def api_branches():
    return jsonify(BRANCHES)


@app.route("/api/routes")
def api_routes():
    origin = request.args.get("origin")
    destination = request.args.get("destination")
    if not origin or not destination:
        return jsonify([])
    return jsonify(get_routes(origin, destination))


@app.route("/api/mileage-table")
def api_mileage_table():
    """Return the full mileage table for the reference card."""
    table = []
    for (origin, dest), routes in MILEAGE_TABLE.items():
        for r in routes:
            table.append({
                "origin": origin,
                "destination": dest,
                "route": r["route"],
                "miles": r["miles"],
                "reimbursement": calculate_reimbursement(r["miles"]),
            })
    return jsonify(table)


@app.route("/api/rate")
def api_rate():
    return jsonify({"rate": REIMBURSEMENT_RATE})


# ---------------------------------------------------------------------------
# API — CRUD (all scoped to user_id)
# ---------------------------------------------------------------------------
@app.route("/api/entries")
@require_user
def api_entries():
    """Get entries for a given month/year, scoped to the current user."""
    user_id = get_user_id()

    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    if not year or not month:
        return jsonify({"error": "year and month are required"}), 400

    conn = get_db()
    try:
        rows = conn.execute(
            """SELECT * FROM entries
               WHERE user_id = ? AND year = ? AND month = ?
               ORDER BY day, id""",
            (user_id, year, month),
        ).fetchall()
    finally:
        conn.close()

    entries = [dict(r) for r in rows]
    total_miles = round(sum(e["miles"] for e in entries), 2)
    total_reimbursement = round(
        sum(e["reimbursement_amount"] for e in entries), 2
    )

    return jsonify({
        "entries": entries,
        "summary": {
            "total_entries": len(entries),
            "total_miles": total_miles,
            "total_reimbursement": total_reimbursement,
        },
    })


@app.route("/api/entries", methods=["POST"])
@require_user
def api_create_entry():
    """Create a new mileage entry, optionally with a return trip."""
    user_id = get_user_id()

    data = request.get_json()
    errors = _validate_entry(data)
    if errors:
        return jsonify({"errors": errors}), 400

    values, err = _build_entry_values(data, user_id)
    if err:
        return jsonify({"errors": [err]}), 400

    conn = get_db()
    try:
        created = []

        # --- Main entry ---
        cursor = conn.execute(INSERT_SQL, values)
        main_id = cursor.lastrowid
        main_row = conn.execute(
            "SELECT * FROM entries WHERE id = ?", (main_id,)
        ).fetchone()
        created.append(dict(main_row))

        # --- Return trip (if requested) ---
        if data.get("include_return"):
            return_routes = get_routes(
                data["destination_branch"], data["origin_branch"]
            )
            if return_routes:
                return_route = next(
                    (r for r in return_routes
                     if r["route"] == data["route_name"]),
                    return_routes[0],
                )

                ret_values, ret_err = _build_entry_values(
                    data, user_id,
                    override_origin=data["destination_branch"],
                    override_dest=data["origin_branch"],
                    override_route=return_route["route"],
                )
                if ret_values and not ret_err:
                    cursor2 = conn.execute(INSERT_SQL, ret_values)
                    ret_id = cursor2.lastrowid
                    ret_row = conn.execute(
                        "SELECT * FROM entries WHERE id = ?", (ret_id,)
                    ).fetchone()
                    created.append(dict(ret_row))

        conn.commit()
    finally:
        conn.close()

    return jsonify({"entries": created, "count": len(created)}), 201


@app.route("/api/entries/<int:entry_id>", methods=["PUT"])
@require_user
def api_update_entry(entry_id):
    """Update an existing entry (must belong to the current user)."""
    user_id = get_user_id()

    data = request.get_json()
    errors = _validate_entry(data)
    if errors:
        return jsonify({"errors": errors}), 400

    miles = get_route_miles(
        data["origin_branch"], data["destination_branch"], data["route_name"]
    )
    if miles is None:
        return jsonify({"errors": ["Invalid route selection"]}), 400

    reimbursement = calculate_reimbursement(miles)
    year = int(data["year"])
    month = int(data["month"])
    day = int(data["day"])
    date_str = f"{year}-{month:02d}-{day:02d}"
    purpose = data.get("business_purpose", "").strip()
    if not purpose:
        purpose = (
            f"Mileage: {data['origin_branch']} to "
            f"{data['destination_branch']} via {data['route_name']}"
        )
    notes = data.get("notes", "").strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_db()
    try:
        conn.execute(
            """UPDATE entries
               SET year = ?, month = ?, day = ?, date = ?, origin_branch = ?,
                   destination_branch = ?, route_name = ?, miles = ?,
                   reimbursement_amount = ?, business_purpose = ?, notes = ?,
                   updated_at = ?
               WHERE id = ? AND user_id = ?""",
            (year, month, day, date_str, data["origin_branch"],
             data["destination_branch"], data["route_name"], miles,
             reimbursement, purpose, notes, now, entry_id, user_id),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM entries WHERE id = ? AND user_id = ?",
            (entry_id, user_id),
        ).fetchone()
    finally:
        conn.close()

    if row is None:
        return jsonify({"error": "Entry not found"}), 404
    return jsonify(dict(row))


@app.route("/api/entries/<int:entry_id>", methods=["DELETE"])
@require_user
def api_delete_entry(entry_id):
    """Delete a single entry (must belong to the current user)."""
    user_id = get_user_id()

    conn = get_db()
    try:
        row = conn.execute(
            "SELECT id FROM entries WHERE id = ? AND user_id = ?",
            (entry_id, user_id),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Entry not found"}), 404
        conn.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
        conn.commit()
    finally:
        conn.close()

    return jsonify({"deleted": entry_id})


@app.route("/api/entries/clear", methods=["POST"])
@require_user
def api_clear_month():
    """Delete all entries for a given month/year for the current user."""
    user_id = get_user_id()

    data = request.get_json()
    year = data.get("year")
    month = data.get("month")
    if not year or not month:
        return jsonify({"error": "year and month are required"}), 400

    conn = get_db()
    try:
        result = conn.execute(
            "DELETE FROM entries WHERE user_id = ? AND year = ? AND month = ?",
            (user_id, year, month),
        )
        conn.commit()
        deleted = result.rowcount
    finally:
        conn.close()

    return jsonify({"deleted_count": deleted})


# ---------------------------------------------------------------------------
# API — Export (scoped to user_id)
# ---------------------------------------------------------------------------
@app.route("/api/export/csv")
@require_user
def api_export_csv():
    """Export entries for a month as CSV."""
    user_id = get_user_id()
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    if not year or not month:
        return jsonify({"error": "year and month are required"}), 400

    conn = get_db()
    try:
        rows = conn.execute(
            """SELECT * FROM entries
               WHERE user_id = ? AND year = ? AND month = ?
               ORDER BY day, id""",
            (user_id, year, month),
        ).fetchall()
    finally:
        conn.close()

    month_name = calendar.month_name[month]
    filename = f"mileage_{month_name}_{year}.csv"

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Date", "Day", "From", "To", "Route", "Miles",
        "Reimbursement ($)", "Business Purpose", "Notes",
    ])
    total_miles = 0
    total_reimb = 0
    for r in rows:
        writer.writerow([
            r["date"], r["day"], r["origin_branch"],
            r["destination_branch"], r["route_name"], r["miles"],
            f"{r['reimbursement_amount']:.2f}",
            r["business_purpose"], r["notes"],
        ])
        total_miles += r["miles"]
        total_reimb += r["reimbursement_amount"]
    writer.writerow([])
    writer.writerow([
        "", "", "", "", "TOTALS",
        f"{total_miles:.2f}", f"{total_reimb:.2f}", "", "",
    ])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/export/excel")
@require_user
def api_export_excel():
    """Export entries using the official expense reimbursement template.

    The template has 10 pre-formatted data rows (9–18) with formulas:
      - Column E: ``=D{n}*0.725`` (mileage reimbursement)
      - Columns F–I: other expense categories (default 0)
      - Column J: ``=SUM(E{n}:I{n})`` (row total)
      - Row 19: column totals via ``=SUM()``

    This function fills columns A–D with tracker data and preserves all
    formulas and formatting.  If entries exceed 10, extra rows are
    inserted with cloned formatting and formulas.
    """
    user_id = get_user_id()
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    if not year or not month:
        return jsonify({"error": "year and month are required"}), 400

    conn = get_db()
    try:
        rows = conn.execute(
            """SELECT * FROM entries
               WHERE user_id = ? AND year = ? AND month = ?
               ORDER BY day, id""",
            (user_id, year, month),
        ).fetchall()

        user_row = conn.execute(
            "SELECT display_name FROM users WHERE id = ?", (user_id,)
        ).fetchone()
    finally:
        conn.close()

    display_name = user_row["display_name"] if user_row else ""
    month_name = calendar.month_name[month]
    filename = f"expense_reimbursement_{month_name}_{year}.xlsx"

    # --- Load the official template ---
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Template layout constants
    FIRST_DATA_ROW = 9      # First entry row
    TEMPLATE_ROWS = 10      # Template has rows 9–18 for data
    TOTALS_ROW = 19         # Row with SUM formulas

    num_entries = len(rows)
    extra_rows = max(0, num_entries - TEMPLATE_ROWS)

    # --- Insert extra rows if more than 10 entries ---
    if extra_rows > 0:
        ws.insert_rows(FIRST_DATA_ROW + TEMPLATE_ROWS, amount=extra_rows)

        source_row = FIRST_DATA_ROW + TEMPLATE_ROWS - 1  # row 18
        for offset in range(extra_rows):
            new_row = FIRST_DATA_ROW + TEMPLATE_ROWS + offset
            for col in range(1, 11):  # A–J
                src_cell = ws.cell(row=source_row, column=col)
                dst_cell = ws.cell(row=new_row, column=col)
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy.copy(src_cell.alignment)
            # Formulas for the new row
            ws.cell(row=new_row, column=5).value = f"=D{new_row}*0.725"
            ws.cell(row=new_row, column=6).value = 0   # Fuel
            ws.cell(row=new_row, column=7).value = 0   # Meals/Ent
            ws.cell(row=new_row, column=8).value = 0   # Phone
            ws.cell(row=new_row, column=9).value = 0   # Other
            ws.cell(row=new_row, column=10).value = \
                f"=SUM(E{new_row}:I{new_row})"

        # Update the TOTALS row formulas
        actual_totals_row = TOTALS_ROW + extra_rows
        last_data_row = FIRST_DATA_ROW + num_entries - 1
        for col_letter in ("D", "E", "F", "G", "H", "I"):
            ws[f"{col_letter}{actual_totals_row}"] = \
                f"=SUM({col_letter}{FIRST_DATA_ROW}:" \
                f"{col_letter}{last_data_row})"
        ws[f"J{actual_totals_row}"] = \
            f"=SUM(E{actual_totals_row}:I{actual_totals_row})"

    # --- Fill header fields ---
    ws["B5"] = display_name   # Name

    # --- Fill data rows ---
    for i, r in enumerate(rows):
        row_num = FIRST_DATA_ROW + i

        # A: Date (as a proper datetime so the m/d/yy format applies)
        ws.cell(row=row_num, column=1).value = datetime(
            int(r["year"]), int(r["month"]), int(r["day"])
        )

        # B: Description
        ws.cell(row=row_num, column=2).value = (
            f"{r['origin_branch']} to "
            f"{r['destination_branch']} via {r['route_name']}"
        )

        # C: Business Purpose
        ws.cell(row=row_num, column=3).value = r["business_purpose"]

        # D: Miles (numeric — the E column formula multiplies this)
        ws.cell(row=row_num, column=4).value = r["miles"]

    # --- Save to buffer ---
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        },
    )


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True, port=5000)
